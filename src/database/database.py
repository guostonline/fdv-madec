import pymongo
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import logging
from src.vendeurs_phone import vendeur_number_phone

class SaveSuivi:
    
    def __init__( self, excel,date:str) -> None:
        self.vendeurs=  list(vendeur_number_phone.keys())
        self.excel = excel
        self.date=date

    def transform_df(self, vendeur):
        wb = load_workbook(self.excel)
        sheet_ranges = wb["AGADIR"]
        my_date = {}

        for i in range(9, sheet_ranges.max_row):
            if sheet_ranges[f"C{i}"].value == vendeur:
                my_date.update(
                    {sheet_ranges[f"D{i}"].value: sheet_ranges[f"E{i}"].value}
                )
                
        return my_date
               

    def send_data(self, vendeur: str, data: dict):
        myclient = pymongo.MongoClient("mongodb://localhost:27017/")
        database = myclient["madec"]
        my_collection = database["suivi_journalier"]
        
        data_send = {"vendeur": vendeur, "date": self.date, "data": data}
        test = my_collection.insert_one(data_send)
        
    def send_all_vendeurs(self):
        for vendeur in self.vendeurs:
          
           self.send_data(vendeur, self.transform_df(vendeur))
        logging("All vendeurs send successfully")

    def find_vendeur(self,vendeur):
        myclient = pymongo.MongoClient("mongodb://localhost:27017/")
        database = myclient["madec"]
        my_collection = database["suivi_journalier"]
        for x in my_collection.find({},{"vendeur" :vendeur}):
            print(x)